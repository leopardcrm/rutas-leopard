import openpyxl
import urllib.request
import re
import json
import time
import os

def extract_coords(url):
    # Regex to find coordinates like -17.749642,-63.125733
    # Try finding in /place/lat,lng/
    match = re.search(r'/place/(-?\d+\.\d+),(-?\d+\.\d+)', url)
    if match:
        return float(match.group(1)), float(match.group(2))
    
    # Try finding in /search/lat,lng/
    match = re.search(r'/search/(-?\d+\.\d+),(-?\d+\.\d+)', url)
    if match:
        return float(match.group(1)), float(match.group(2))
        
    # Try finding in /@lat,lng,
    match = re.search(r'@(-?\d+\.\d+),(-?\d+\.\d+)', url)
    if match:
        return float(match.group(1)), float(match.group(2))
        
    # Generic lat,lng check in path
    match = re.search(r'/(-?\d+\.\d+),(-?\d+\.\d+)', url)
    if match:
        return float(match.group(1)), float(match.group(2))
        
    return None

def resolve_url(short_url):
    try:
        req = urllib.request.Request(
            short_url, 
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        )
        with urllib.request.urlopen(req, timeout=10) as response:
            return response.geturl()
    except Exception as e:
        print(f"Error resolving {short_url}: {e}")
        return None

def main():
    wb = openpyxl.load_workbook('Leopard S. R. L.  (respuestas).xlsx')
    sheet = wb.active
    
    clients = []
    
    # Column mapping:
    # 1: Marca temporal (Timestamp)
    # 2: Nombre del cliente (Client Name)
    # 3: Celular (Phone)
    # 4: Nombre de la Tienda (Business Name)
    # 5: Maps (Google Maps URL)
    # 6: Foto (Photo URL)
    
    total_rows = sheet.max_row
    print(f"Total rows to process: {total_rows - 1}")
    
    for row_idx in range(2, total_rows + 1):
        timestamp = sheet.cell(row=row_idx, column=1).value
        name = sheet.cell(row=row_idx, column=2).value
        phone = sheet.cell(row=row_idx, column=3).value
        shop_name = sheet.cell(row=row_idx, column=4).value
        maps_url = sheet.cell(row=row_idx, column=5).value
        photo_url = sheet.cell(row=row_idx, column=6).value
        
        if not name or not maps_url:
            continue
            
        print(f"[{row_idx-1}/{total_rows-1}] Processing: {name} - {shop_name}")
        
        # Clean and format phone
        phone_str = ""
        if phone:
            try:
                phone_str = str(int(float(phone)))
            except:
                phone_str = str(phone)
                
        # Clean strings
        name = str(name).strip() if name else ""
        shop_name = str(shop_name).strip() if shop_name else ""
        maps_url = str(maps_url).strip() if maps_url else ""
        photo_url = str(photo_url).strip() if photo_url else ""
        
        lat, lng = None, None
        resolved_url = resolve_url(maps_url)
        if resolved_url:
            coords = extract_coords(resolved_url)
            if coords:
                lat, lng = coords
                print(f"  Found coordinates: {lat}, {lng}")
            else:
                print(f"  Could not extract coordinates from resolved URL: {resolved_url}")
        else:
            print(f"  Could not resolve short URL: {maps_url}")
            
        # If coordinates not found, assign default coordinate (Santa Cruz, Bolivia area, where Leopard SRL seems to operate based on phone codes)
        # Lat: -17.786, Lng: -63.181 (approx center)
        is_approximate = False
        if lat is None or lng is None:
            lat = -17.7833
            lng = -63.1821
            is_approximate = True
            print("  Warning: Using default coordinate for Santa Cruz, Bolivia.")
            
        client_data = {
            "id": row_idx - 1,
            "timestamp": str(timestamp) if timestamp else "",
            "client_name": name,
            "phone": phone_str,
            "shop_name": shop_name,
            "maps_url": maps_url,
            "photo_url": photo_url,
            "latitude": lat,
            "longitude": lng,
            "is_approximate": is_approximate,
            "seller_id": ((row_idx - 2) % 6) + 1, # Distribute among the 6 sellers for demo purposes
            "status": "Pendiente"
        }
        clients.append(client_data)
        
        # Rate limit friendly
        time.sleep(0.5)
        
    # Write to JSON
    with open('clients_imported.json', 'w', encoding='utf-8') as f:
        json.dump(clients, f, ensure_ascii=False, indent=2)
        
    print(f"Finished! Imported {len(clients)} clients into clients_imported.json")

if __name__ == "__main__":
    main()
